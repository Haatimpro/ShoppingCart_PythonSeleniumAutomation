import openpyxl
def excel_update_data(book_path,searchterm,column_name,new_value):
    #edit the excel with updated value
    #------------------------------------------------------------------------------
    #book_path = "C:\\Users\\Sandesh\\Downloads\\download.xlsx"

    book = openpyxl.load_workbook(book_path)
    sheet = book.active
    maximumRow = sheet.max_row
    maximumColumn = sheet.max_column
    Dict = {}

    #2nd build the logic to identi the price column, as we will have column headers in 1st row, harcode row = 1, and iterate column with i
    #When I matches for column name = "Price" then add it to dictionary
    for i in range(1,maximumColumn+1):
        if sheet.cell(row=1,column=i) == column_name:
            Dict["col"] = i

    #find the apple fruit name row
    for i in range(i,maximumRow+1):
        for j in range(1,maximumColumn+1):
            if sheet.cell(row = i, column = j).value == searchterm:
                Dict["row"] = i


    sheet.cell(row=Dict["row"],column=Dict["col"]).value = new_value
    book.save(book_path)