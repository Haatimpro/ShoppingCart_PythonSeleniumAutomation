class HomePageTestData:

        Homepage_TestData = [{"email" : "sandeshtolar@gamil.com", "password" : "123asd"}, {"email" : "yoyo@gamil.com", "password" : "sdjfur454"}]

        @staticmethod
        def getTestData(test_case_name):
                Dict = {}
                import openpyxl
                book = openpyxl.load_workbook("D:\\Selenium_Training\\notes\\ExcelPythonDemo.xlsx")
                sheet = book.active
                maxrow = sheet.max_row
                maxcolumn = sheet.max_column
                for i in range(1, maxrow + 1):
                        if sheet.cell(row=i, column=1).value == test_case_name:
                                # start column iteration
                                for j in range(2, maxcolumn + 1):
                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                return [Dict]