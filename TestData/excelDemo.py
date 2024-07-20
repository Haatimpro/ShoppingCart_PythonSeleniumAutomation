import openpyxl
book = openpyxl.load_workbook("D:\\Selenium_Training\\notes\\ExcelPythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
# sheet.cell(row=5,column=5).value = "sandesh"
# print(sheet.cell(row=5,column=5).value)
maxrow = sheet.max_row
print(f"maximum row is:  {maxrow}")
maxcolumn=sheet.max_column
print(f"maximum column is: {maxcolumn}")
Dict = {}

#print all cell values in the sheet
#start with row iteration
for i in range(1,maxrow+1):
    if sheet.cell(row=i,column=1).value == "testcase2":
        #start column iteration
        for j in range(2,maxcolumn+1):
            Dict[sheet.cell(row=1,column=j).value]  = sheet.cell(row=i,column=j).value
print(Dict)