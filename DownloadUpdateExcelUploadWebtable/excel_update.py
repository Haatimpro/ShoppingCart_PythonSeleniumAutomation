import openpyxl

class Excel:

    def test_ExcelUpdate(self):
        #2.Access the downloaded excel and update Banana price to 1000
        #--------------------------------------------------------------------------------------
        book = openpyxl.load_workbook("C:\\Users\\Sandesh\\Downloads\\download.xlsx")
        sheet = book.active
        maximunrow= sheet.max_row
        maximumcolumn = sheet.max_column
        Dict={}

        #find price column number, as column header is always in 1st row, hardcode row to 1
        for j in range(1,maximumcolumn+1):
            if sheet.cell(row=1,column=j).value == "price":
                Dict["col"] = j

        #find banana fruit row number, from complete data set
        for i in range(1,maximunrow+1):
            for j in range(1,maximumcolumn+1):
                if sheet.cell(row=i,column=j).value == 'Banana':
                    Dict["row"] = i
        #update the required value
        sheet.cell(row=Dict["row"], column=Dict["col"]).value = 1000 #error encountered key error: "col" because in 1st loop
        # key "col" didnot initialized as "Price" column not found because in excel colum is "price" not "Price"

        #save the changes
        book.save("C:\\Users\\Sandesh\\Downloads\\download.xlsx")
